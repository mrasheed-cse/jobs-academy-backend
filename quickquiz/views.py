import uuid
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from rest_framework.permissions import AllowAny
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework import status, viewsets, generics
from django.core.files.base import ContentFile
from django.db import transaction
from openpyxl.utils import get_column_letter
from rest_framework.permissions import IsAuthenticated
import random
from django.db.models import Sum
from datetime import datetime, timedelta
from rest_framework import status
from rest_framework.views import APIView
from .models import *
from .serializers import *
from django.db.models import Q

from django.shortcuts import get_object_or_404
from django.db import IntegrityError
from decimal import Decimal
from django.contrib.auth import get_user_model
User = get_user_model()
from rest_framework.generics import ListAPIView
from django.db.models import Count, Avg, Sum
from rest_framework.permissions import IsAdminUser, AllowAny
from quiz.permissions import *
from rest_framework.pagination import PageNumberPagination
from django.utils.dateparse import parse_date
from datetime import datetime, time

from django.db.models import Sum, Q, Window
from django.db.models.functions import Rank

class SubjectViewSet(viewsets.ModelViewSet):
    # queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    
    def get_queryset(self):
        return (
            Subject.objects.annotate(question_count=Count("practicequestion"))
            .filter(question_count__gte=10)  # only subjects with at least 10 questions
        )
    
# Start a new Practice Session
class StartPracticeSessionView(APIView):
    # authentication_classes = [AllowInactiveUserJWTAuthentication]
    permission_classes = [AllowAny]
    def post(self, request):
        user = request.user  # Assuming the user is authenticated
        subject_id = request.data.get('subject_id')
        
        if not subject_id:
            return Response({'error': 'subject_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            subject = Subject.objects.get(id=subject_id)
            
        except Subject.DoesNotExist:
            return Response({'error': 'Subject not found.'}, status=status.HTTP_404_NOT_FOUND)

        questions = list(PracticeQuestion.objects.filter(subject=subject))
        
        if len(questions) < 10:
            return Response({'error': 'Not enough questions available for this subject.'}, status=status.HTTP_400_BAD_REQUEST)

        selected_questions = random.sample(questions, 10)

        question_data = PracticeQuestionSerializer(selected_questions, many=True).data

        return Response({
            'questions': question_data
        }, status=status.HTTP_201_CREATED)


# Submit Answers and Calculate Score
class SubmitPracticeSessionView(APIView):
    permission_classes = [AllowAny]  # Allow both authenticated and unauthenticated

    def post(self, request):
        user = None
        phone_number = request.data.get('phone_number')
        username = request.data.get('username')
        # print(request.data)
        if request.user and request.user.is_authenticated:
            user = request.user
        elif phone_number:
            # Try to find an existing UserPoints by phone_number
            try:
                user_points = UserPoints.objects.get(phone_number=phone_number)
                username = user_points.username  # Overwrite username if already stored
            except UserPoints.DoesNotExist:
                user_points = UserPoints.objects.create(username=username, phone_number=phone_number)
        else:
            return Response({'error': 'Authentication or phone number is required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Create the session
        session = PracticeSession.objects.create(user=user, username=username, phone_number=phone_number)

        # Handle duration
        duration = request.data.get('duration')
        if duration is not None:
            try:
                session.duration = timedelta(minutes=int(duration))
            except (ValueError, TypeError):
                return Response({'error': 'Invalid duration format.'}, status=status.HTTP_400_BAD_REQUEST)

        answers = request.data.get('answers', [])  # [{'question_id': ..., 'option_id': ...}]
        score = 0
        correct_answers = 0
        total_questions = len(answers)

        for question in answers:
            selected_option_id = question.get('selected_option_id')
            
            if selected_option_id:
                try:
                    selected_option = PracticeOption.objects.get(id=int(selected_option_id))
                    if selected_option.is_correct:
                        score += 1
                        correct_answers += 1
                except PracticeOption.DoesNotExist:
                    continue

        wrong_answers = total_questions - correct_answers
        percentage_score = round((correct_answers / total_questions) * 100, 2) if total_questions > 0 else 0

        # Save score
        session.score = score
        session.save()

        # Update or create UserPoints
        if user:
            user_points, _ = UserPoints.objects.get_or_create(user=user)
        elif phone_number:
            user_points, _ = UserPoints.objects.get_or_create(phone_number=phone_number, defaults={'username': username})

        user_points.add_points(score)

        return Response({
            'score': score,
            'correct_answers': correct_answers,
            'wrong_answers': wrong_answers,
            'percentage': percentage_score,
            'duration_in_minutes': round(session.duration.total_seconds() / 60) if session.duration else 0,
        }, status=status.HTTP_200_OK)

# View for Leaderboard - Display Top 10 Users with highest points
class PracticeLeaderboardAPIView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        # Fetch top 10 users based on points
        # top_users_points = (
        #     UserPoints.objects.select_related('user')
        #     .order_by('-points')[:10]
        # )
        top_users_points = (
            UserPoints.objects.select_related('user')
            .order_by('-points')
        )

        top_data = []
        for up in top_users_points:
            user = up.user
            profile_image = None
            if user and hasattr(user, 'profile_picture') and user.profile_picture:
                profile_image = request.build_absolute_uri(user.profile_picture.url)

            attempts = PracticeSession.objects.filter(user=user).count() if user else PracticeSession.objects.filter(phone_number=up.phone_number).count()

            top_data.append({
                'id': user.id if user else None,
                'username': user.username if user else up.username,
                'points': up.points,
                'attempts': attempts,
                'profile_image': profile_image,
            })

        current_user_data = None

        # Check if user is authenticated
        if request.user.is_authenticated:
            user = request.user
            user_points = UserPoints.objects.filter(user=user).first()
            if user_points:
                rank_list = list(UserPoints.objects.order_by('-points').values_list('user_id', flat=True))
                try:
                    rank = rank_list.index(user.id) + 1
                except ValueError:
                    rank = None

                profile_image = request.build_absolute_uri(user.profile_picture.url) if user.profile_picture else None

                current_user_data = {
                    'id': user.id,
                    'username': user.username,
                    'points': user_points.points,
                    'rank': rank,
                    'attempts': PracticeSession.objects.filter(user=user).count(),
                    'profile_image': profile_image,
                }

        # If not authenticated, check for phone_number in query params
        elif phone_number := request.query_params.get('phone_number'):
            user_points = UserPoints.objects.filter(phone_number=phone_number).first()
            if user_points:
                ranks = list(UserPoints.objects.order_by('-points').values_list('phone_number', flat=True))
                try:
                    rank = ranks.index(phone_number) + 1
                except ValueError:
                    rank = None

                current_user_data = {
                    'id': None,
                    'username': user_points.username or "Guest",
                    'points': user_points.points,
                    'rank': rank,
                    'attempts': PracticeSession.objects.filter(phone_number=phone_number).count(),
                    'profile_image': None,
                }

        return Response({
            'top_10': top_data,
            'me': current_user_data  # This will be None if no user info was found
        })




class PracticeQuestionUploadView(APIView):
    permission_classes = [IsTeacherOrAdmin]
    parser_classes = [MultiPartParser]

    # Optional: You may need to implement this method depending on your needs
    def extract_images(self, workbook):
        # This should return a dictionary mapping cell refs to image data
        return {}

    def get_image_data_from_map(self, img_obj):
        return img_obj

    def save_image_to_field(self, image_data, filename):
        from django.core.files.base import ContentFile
        return ContentFile(image_data, name=filename)


    def safe_str(self, val):
        import pandas as pd
        if pd.isna(val):
            return ""
        return str(val).strip()

    

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file or not file.name.endswith('.xlsx'):
            return Response({'error': 'Please upload a valid .xlsx Excel file.'}, status=400)

        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            df = pd.DataFrame(ws.values)

            # Normalize headers
            df.columns = df.iloc[0].astype(str).str.strip().str.lower()
            df = df[1:].reset_index(drop=True)

            image_map = self.extract_images(wb)

            created_count, skipped_count = 0, 0

            # Label Mappings
            LABEL_MAP = {
                'option1': 'a', 'option2': 'b', 'option3': 'c', 'option4': 'd',
                'a': 'a', 'b': 'b', 'c': 'c', 'd': 'd',
                'A': 'a', 'B': 'b', 'C': 'c', 'D': 'd',
                'ক': 'a', 'খ': 'b', 'গ': 'c', 'ঘ': 'd'
            }
            OPTION_LABELS = ['a', 'b', 'c', 'd']

            with transaction.atomic():
                last_subject = None  # Track last non-empty subject

                for index, row in df.iterrows():
                    excel_row_num = index + 2

                    question_text = self.safe_str(row.get('question'))
                    q_cell = f"{get_column_letter(df.columns.get_loc('question') + 1)}{excel_row_num}"
                    question_image = self.get_image_data_from_map(image_map.get(q_cell))

                    if not question_text and not question_image:
                        # skipped_count += 1
                        continue

                    # Handle subject
                    subject_name = self.safe_str(row.get('subject'))

                    subject = None
                    if subject_name:
                        subject, _ = Subject.objects.get_or_create(name=subject_name)
                        last_subject = subject  # update last_subject if new one found
                    else:
                        subject = last_subject  # use previous subject if blank

                    
                    # ✅ DUPLICATE CHECK
                    duplicate_exists = PracticeQuestion.objects.filter(
                        text__iexact=question_text.strip(),
                        subject=subject
                    ).exists()

                    if duplicate_exists:
                        skipped_count += 1
                        continue
                    
                    
                    # Create question
                    question = PracticeQuestion.objects.create(
                        text=question_text,
                        marks=1,
                        subject=subject
                    )

                    if question_image:
                        q_filename = f"q_img_{question.id}_{uuid.uuid4().hex[:8]}.png"
                        q_file = self.save_image_to_field(question_image, q_filename)
                        if q_file:
                            question.image.save(q_file.name, q_file, save=True)

                    correct = self.safe_str(row.get("answer")).lower()
                    standard_correct = LABEL_MAP.get(correct, correct)

                    # Map options
                    option_columns = {}
                    for col in df.columns:
                        key = str(col).strip()
                        label = LABEL_MAP.get(key)
                        if label and label not in option_columns:
                            option_columns[label] = key

                    for label in OPTION_LABELS:
                        col_key = option_columns.get(label)
                        if not col_key:
                            continue

                        col_index = df.columns.get_loc(col_key)
                        cell_ref = f"{get_column_letter(col_index + 1)}{excel_row_num}"
                        opt_image = self.get_image_data_from_map(image_map.get(cell_ref))
                        opt_text = self.safe_str(row.get(col_key))


                        is_correct = (standard_correct == label or standard_correct == opt_text.lower())

                        if opt_image:
                            opt_filename = f"opt_img_{question.id}_{uuid.uuid4().hex[:8]}.png"
                            opt_file = self.save_image_to_field(opt_image, opt_filename)
                            if opt_file:
                                PracticeOption.objects.create(
                                    question=question,
                                    text=None,
                                    image=opt_file,
                                    is_correct=is_correct
                                )
                        elif opt_text:
                            PracticeOption.objects.create(
                                question=question,
                                text=opt_text,
                                is_correct=is_correct
                            )

                    created_count += 1


            return Response({
                "message": "Upload complete",
                "questions_created": created_count,
                "questions_skipped": skipped_count
            })

        except Exception as e:
            return Response({'error': str(e)}, status=500)
        
        


    def extract_images(self, workbook):
        """Returns a map of Excel cell positions to image binary data"""
        image_map = {}
        for sheet in workbook.worksheets:
            for image in getattr(sheet, '_images', []):
                if hasattr(image, 'anchor'):
                    cell = image.anchor._from
                    cell_ref = f"{get_column_letter(cell.col + 1)}{cell.row + 1}"
                    with BytesIO() as output:
                        image.ref.save(output, format='PNG')
                        image_map[cell_ref] = output.getvalue()
        return image_map

    # def get_image_data_from_map(self, image_data):
    #     if image_data:
    #         return image_data
    #     return None

    # def save_image_to_field(self, image_data, filename):
    #     return ContentFile(image_data, name=filename)
    
from django.utils.timezone import localtime
from django.db.models import OuterRef, Subquery
from datetime import datetime, date
# class DailyTopScorerAPIView(APIView):
#     permission_classes = [AllowAny]

#     def get(self, request):
#         selected_date = request.GET.get("date")

#         # ✅ Validate date input
#         if not selected_date:
#             selected_date = localtime().date()
#         else:
#             try:
#                 selected_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
#             except ValueError:
#                 return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)

#         # ✅ Get all practice sessions of the day
#         sessions = PracticeSession.objects.filter(
#             created_at__date=selected_date
#         ).select_related("user")

#         if not sessions.exists():
#             return Response({
#                 "date": selected_date,
#                 "total_attempts": 0,
#                 "unique_users": 0,
#                 "leaderboard": [],
#                 "message": "No attempts found for this date."
#             })

#         # ✅ Aggregate total score per user/guest
#         leaderboard_data = {}
#         for s in sessions:
#             if s.user:  # authenticated user
#                 key = f"user_{s.user_id}"
#                 username = s.user.username
#                 phone_number = s.user.phone_number
#             else:  # guest user (unauthorized)
#                 key = f"guest_{s.username}_{s.phone_number}"
#                 username = s.username or "Guest"
#                 phone_number = s.phone_number or "N/A"

#             if key not in leaderboard_data:
#                 leaderboard_data[key] = {
#                     "user_id": s.user_id,
#                     "username": username,
#                     "phone_number": phone_number,
#                     "total_score": 0,
#                     "attempts": 0
#                 }

#             leaderboard_data[key]["total_score"] += s.score
#             leaderboard_data[key]["attempts"] += 1

#         # ✅ Convert to sorted list
#         leaderboard = sorted(
#             leaderboard_data.values(),
#             key=lambda x: x["total_score"],
#             reverse=True
#         )

#         return Response({
#             "date": selected_date,
#             "total_attempts": sessions.count(),
#             "unique_users": len(leaderboard_data),
#             "leaderboard": leaderboard,
#         })



class DailyTopScorerAPIView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        date = request.GET.get("date")
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        print("date is here", date, start_date, end_date)
        
        # ✅ Handle date parsing and default values
        try:
            if start_date and end_date:
                start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            elif date:
                start_date = end_date = date
            else:
                # Default: today's date range
                start_date = end_date = localtime().date()
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD"},
                status=400
            )

        # ✅ Query all sessions within the date range
        sessions = PracticeSession.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).select_related("user")
        

        if not sessions.exists():
            return Response({
                "date_range": f"{start_date} to {end_date}",
                "total_attempts": 0,
                "unique_users": 0,
                "leaderboard": [],
                "message": "No attempts found in this date range."
            })

        # ✅ Aggregate total score per user (including guests)
        leaderboard_data = {}
        
        for s in sessions:
            if s.user:  # Authenticated user
                key = f"{s.user_id}"
                username = s.user.username
                phone_number = s.user.phone_number
            else:  # Guest user
                key = f"{s.username}_{s.phone_number}"
                
                username = s.username or "Guest"
                phone_number = s.phone_number or "N/A"

            if key not in leaderboard_data:
                leaderboard_data[key] = {
                    "user_id": s.user_id,
                    "username": username,
                    "phone_number": phone_number,
                    "total_score": 0,
                    "attempts": 0
                }

            leaderboard_data[key]["total_score"] += s.score
            leaderboard_data[key]["attempts"] += 1

        # ✅ Sort leaderboard by total_score descending
        leaderboard = sorted(
            leaderboard_data.values(),
            key=lambda x: x["total_score"],
            reverse=True
        )
        print(leaderboard)
        # ✅ Response
        return Response({
            "date_range": f"{start_date} to {end_date}",
            "total_attempts": sessions.count(),
            "unique_users": len(leaderboard_data),
            "leaderboard": leaderboard,
        })
        
class AdminAnalyticsAPIView(APIView):
    """
    GET /api/admin/analytics/
    Returns summarized analytics data for admin dashboard.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = timezone.now().date()
        last_7_days = today - timedelta(days=7)

        # ✅ General stats
        total_users = User.objects.count()
        total_sessions = PracticeSession.objects.count()
        total_questions = PracticeQuestion.objects.count()
        total_subjects = Subject.objects.count()
        total_points = UserPoints.objects.aggregate(total=Sum("points"))["total"] or 0

        # ✅ Top 5 users by total score
        top_users = (
            PracticeSession.objects.values("user__username")
            .annotate(total_score=Sum("score"))
            .order_by("-total_score")[:5]
        )

        # ✅ Top 5 subjects by number of questions
        top_subjects = (
            PracticeQuestion.objects.values("subject__name")
            .annotate(question_count=Count("id"))
            .order_by("-question_count")[:5]
        )

        # ✅ Full table: each subject and how many questions it has
        subject_question_table = (
            Subject.objects.annotate(question_count=Count("practicequestion"))
            .values("name", "question_count")
            .order_by("name")
        )

        # ✅ Daily activity for last 7 days
        daily_activity = (
            PracticeSession.objects.filter(created_at__date__gte=last_7_days)
            .values("created_at__date")
            .annotate(
                session_count=Count("id"),
                avg_score=Avg("score"),
                avg_duration=Avg("duration")
            )
            .order_by("created_at__date")
        )

        # ✅ Latest 10 sessions
        recent_sessions = (
            PracticeSession.objects.select_related("user")
            .order_by("-created_at")[:10]
            .values("user__username", "score", "duration", "created_at")
        )

        # ✅ Build response data
        data = {
            "summary": {
                "total_users": total_users,
                "total_sessions": total_sessions,
                "total_questions": total_questions,
                "total_subjects": total_subjects,
                "total_points": total_points,
            },
            "top_users": list(top_users),
            "top_subjects": list(top_subjects),
            "subject_question_table": list(subject_question_table),  # 👈 Added this
            "daily_activity": list(daily_activity),
            "recent_sessions": list(recent_sessions),
        }

        return Response(data)

    


# Reward views section


class RewardDistributionCreateAPIView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        phone_number = request.data.get("phone_number")
        amount_str = request.data.get("amount")
        note = request.data.get("note", "Manual Admin Reward")
        # 👇 NEW: Capture the distribution ID from the frontend
        distribution_id = request.data.get("distribution_id") 
        
        # 1. Validation and Type Conversion
        if not phone_number or not amount_str or not distribution_id:
            return Response(
                {"detail": "Phone number, amount, and distribution ID are required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            reward_amount = Decimal(amount_str).quantize(Decimal('0.01'))
            if reward_amount <= 0:
                 return Response(
                    {"detail": "Amount must be greater than zero."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception:
            return Response(
                {"detail": "Invalid amount format."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 2. Get Specified Distribution
        try:
            # 👇 Fetch the SPECIFIC distribution requested by the admin
            distribution = RewardDistribution.objects.get(id=distribution_id)
        except RewardDistribution.DoesNotExist:
            return Response(
                {"detail": f"Reward distribution with ID {distribution_id} not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # 3. Create or Update UserReward
        try:
            # Check for an existing UserReward for this distribution and phone number
            user_reward, created = UserReward.objects.get_or_create(
                distribution=distribution,
                phone_number=phone_number,
                defaults={
                    'username': 'Admin-Sent', 
                    'total_score': 0,
                    'reward_amount': reward_amount,
                }
            )

            if not created:
                # If the record exists, update the reward amount and note
                original_amount = user_reward.reward_amount
                user_reward.reward_amount = reward_amount
                
                # Update note logic
                current_note = getattr(user_reward, 'note', '') 
                if note != "Manual Admin Reward":
                     # Use 'Manual Adjustment' to distinguish from initial calculated reward
                     user_reward.note = (current_note + f" | Manual Adjustment: ৳{reward_amount} ({note})" ) if current_note else f"Manual Adjustment: ৳{reward_amount} ({note})"
                
                user_reward.save()
            
            # 4. Update Distribution Totals
            if created:
                distribution.total_amount += reward_amount
                distribution.total_users += 1
            else:
                distribution.total_amount += (reward_amount - original_amount)
                
            distribution.save()
            
            return Response({
                "detail": f"Reward of ৳{reward_amount} successfully sent/updated for {phone_number} in distribution ID {distribution_id}.",
                "status": "updated" if not created else "created"
            }, status=status.HTTP_200_OK)

        except IntegrityError:
            return Response(
                {"detail": "A reward record already exists for this user/period and could not be updated."}, 
                status=status.HTTP_409_CONFLICT
            )
        except Exception as e:
            return Response(
                {"detail": f"An error occurred during reward processing: {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class RewardDistributionListAPIView(generics.ListAPIView):
    queryset = RewardDistribution.objects.all().order_by("-distributed_at")
    serializer_class = RewardDistributionSerializer
    permission_classes = [AllowAny]


class UserRewardListAPIView(generics.ListAPIView):
    serializer_class = UserRewardSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        distribution_id = self.kwargs.get("distribution_id")
        return UserReward.objects.filter(distribution_id=distribution_id).order_by("-total_score")


class UserRewardByPhoneAPIView(generics.ListAPIView):
    serializer_class = UserRewardSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        phone_number = self.request.query_params.get("phone_number")
        if not phone_number:
            return UserReward.objects.none()
        return UserReward.objects.filter(phone_number=phone_number).order_by("-distribution__distributed_at")



class UserRewardEfficiencyPagination(PageNumberPagination):
    page_size = 10  # default users per page
    page_size_query_param = 'page_size'
    # max_page_size = 50


class UserRewardEfficiencyView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        distribution = RewardDistribution.objects.order_by('-distributed_at').first()
        if not distribution:
            return Response({"error": "No reward distribution found."}, status=404)

        per_point_value = distribution.per_point_value
        # Assuming UserPoints is the model containing the user's score/points
        user_points = UserPoints.objects.all().select_related('user') # Use select_related for performance

        data = []
        for up in user_points:
            
            # --- START OF REQUIRED LOGIC UPDATE ---
            if up.user:
                # If registered (up.user exists), use phone number from the linked User
                current_phone_number = up.user.phone_number # Assuming the Django User model has a phone_number field
                current_username = up.user.username
            else:
                # If not registered, use phone number directly from UserPoints
                current_phone_number = up.phone_number
                current_username = up.username or "Guest"
            # --- END OF REQUIRED LOGIC UPDATE ---

            # Find the existing reward for this user in this distribution
            user_reward = UserReward.objects.filter(
                phone_number=current_phone_number, 
                distribution=distribution
            ).first()
            
            # Calculate metrics
            rewarded_money = float(user_reward.reward_amount) if user_reward else 0.0
            expected_money = float(up.points) * float(per_point_value)
            difference = expected_money - rewarded_money
            percentage = (rewarded_money / expected_money * 100) if expected_money > 0 else 0

            data.append({
                "username": current_username,
                "phone_number": current_phone_number,
                "points": up.points,
                "rewarded_money": round(rewarded_money, 2),
                "expected_money": round(expected_money, 2),
                "difference": round(difference, 2),
                "percentage": round(percentage, 2),
            })

        # Sort by expected_money descending
        data.sort(key=lambda x: x["difference"], reverse=True)

        # Paginate
        paginator = UserRewardEfficiencyPagination()
        result_page = paginator.paginate_queryset(data, request)
        return paginator.get_paginated_response(result_page)
    
    
class UserRewardStatsAPIView(APIView):
    """
    Provides a personalized reward summary using a phone number.
    """
    permission_classes = [AllowAny] # Removed IsAuthenticated

    def post(self, request): # Changed to POST request
        # 1. Get Phone Number from Request Body
        phone_number = request.data.get("phone_number")
        if not phone_number:
            return Response(
                {"detail": "Phone number is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # NOTE: You'll need logic to fetch the username/user details based on the phone number
        # Assuming a UserProfile or similar model stores this.
        try:
            # Example: Fetch a related user object just for the username (if needed)
            # You might need to adjust this based on your user model structure.
            user_profile = get_object_or_404(User, phone_number=phone_number)
            username = user_profile.username # Or whatever field holds the name
        except Exception:
            # If the phone number isn't registered, we can still proceed with zeros
            username = "Guest User" 

        # 2. Fetch Latest Distribution Data
        latest_distribution = RewardDistribution.objects.order_by('-distributed_at').first()
        
        # Default values
        total_points = Decimal('0.00')
        expected_reward = Decimal('0.00')
        total_rewarded = Decimal('0.00')
        points_to_next_100 = Decimal('0.00') 
        
        if latest_distribution:
            try:
                # Check UserReward using the provided phone_number
                user_reward_record = UserReward.objects.get(
                    distribution=latest_distribution, 
                    phone_number=phone_number
                )
                total_points = user_reward_record.total_score
                expected_reward = user_reward_record.expected_money
                total_rewarded = user_reward_record.rewarded_money
                
                # Recalculate remaining points (same logic as before)
                if total_points > 0:
                    points_modulo_100 = total_points % Decimal('100.00')
                    if points_modulo_100 != 0:
                        points_to_next_100 = Decimal('100.00') - points_modulo_100
                
            except UserReward.DoesNotExist:
                # Phone number is valid but has no reward record in the latest period.
                pass
        
        # 3. Final Output
        return Response({
            "username": username,
            "phone_number": phone_number,
            "total_points": total_points.quantize(Decimal('0.01')),
            "expected_reward_taka": expected_reward.quantize(Decimal('0.01')),
            "total_rewarded_taka": total_rewarded.quantize(Decimal('0.01')),
            "points_to_next_threshold": points_to_next_100.quantize(Decimal('0.01')),
            "current_distribution_period": latest_distribution.start_date.strftime("%b %d") + " - " + latest_distribution.end_date.strftime("%b %d") if latest_distribution else "N/A"
        })
        
# --------------------------------------------------------
# 1. GET PUZZLE LIST
# --------------------------------------------------------
class PuzzleListView(APIView):
    def get(self, request):
        puzzles = WordPuzzle.objects.filter(status="active")
        return Response({
            "type": "success",
            "message": "Puzzle list loaded successfully",
            "data": PuzzleSerializer(puzzles, many=True).data
        })


# --------------------------------------------------------
# 2. GET NEXT PUZZLE WORD (One by One)
# --------------------------------------------------------
class PuzzleWordView(APIView):
    
    def get(self, request, puzzle_id):
        puzzle = get_object_or_404(WordPuzzle, id=puzzle_id)

        words = list(Word.objects.filter(puzzle=puzzle))

        if not words:
            return Response(
                {
                    "type": "error",
                    "message": "No words found",
                    "data": []
                },
                status=404
            )

        # 🔹 Random word every request
        word = random.choice(words)

        return Response({
            "type": "success",
            "message": "Next puzzle word loaded",
            "data": {
                "word_id": word.id,
                "text": word.text,
                "meaning_bn": word.meaning_bn or "",
                "example_en": word.example_en or "",
                "example_bn": word.example_bn or "",
                "hint": word.hint or "",
                "difficulty": word.difficulty,
            }
        })
        
        
class GameListView(ListAPIView):
    # Query only the games you want to show (ordered by newest first)
    queryset = WordPuzzle.objects.all()
    serializer_class = WordPuzzleSerializer
    permission_classes = [AllowAny]        
        
             
class SubmitWordGame(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        score = request.data.get("score")
        puzzle_id = request.data.get("puzzle_id")
        username = request.data.get("username")
        phone = request.data.get("phone_number")

        if score is None or puzzle_id is None:
            return Response(
                {"error": "score and puzzle_id are required"},
                status=400
            )

        try:
            score = int(score)
        except ValueError:
            return Response({"error": "Invalid score"}, status=400)

        puzzle = get_object_or_404(WordPuzzle, id=puzzle_id)

        # ------------------------------------
        # 1. PLAYER (Authenticated / Guest)
        # ------------------------------------
        if request.user.is_authenticated:
            player, _ = Player.objects.get_or_create(
                user=request.user,
                defaults={"username": request.user.username}
            )
        else:
            if not username or not phone:
                return Response(
                    {"error": "username and phone_number are required"},
                    status=400
                )

            player, created = Player.objects.get_or_create(
                phone_number=phone,
                defaults={"username": username}
            )

            if not created and player.username != username:
                player.username = username
                player.save()

        # ------------------------------------
        # 2. CREATE ATTEMPT (IMPORTANT)
        # ------------------------------------
        attempt = WordGameAttempt.objects.create(
            player=player,
            puzzle=puzzle,
            score=score,
            finished_at=timezone.now()
        )

        # ------------------------------------
        # 3. UPDATE SUMMARY SCORE (OPTIONAL)
        # ------------------------------------
        game_score, _ = WordGameScore.objects.get_or_create(player=player)
        game_score.score += score
        game_score.save()

        return Response({
            "message": "Game submitted successfully",
            "attempt_id": attempt.id,
            "player_id": player.id,
            "score": score,
            "total_score": game_score.score,
            "attempt_time": attempt.finished_at
        })

class WordGameLeaderboard(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        top_players = WordGameScore.objects.select_related("player").order_by("-score")[:20]

        data = [
            {
                "username": p.player.username if p.player.username else (p.player.user.username if p.player.user else None),
                "phone": p.player.phone_number,
                "score": p.score
            }
            for p in top_players
        ]

        return Response({"leaderboard": data})



class WordExcelUploadAPIView(APIView):
    permission_classes = [IsAuthenticated]

    REQUIRED_COLUMNS = {"word"}

    def post(self, request):
        serializer = WordExcelUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        puzzle = get_object_or_404(
            WordPuzzle,
            id=serializer.validated_data["puzzle_id"]
        )

        wb = load_workbook(serializer.validated_data["file"])
        sheet = wb.active

        # 🔹 Read header row
        headers = [
            (cell or "").strip().lower()
            for cell in next(
                sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            )
        ]

        # 🔹 Validate required columns
        missing = self.REQUIRED_COLUMNS - set(headers)
        if missing:
            return Response(
                {"error": f"Missing required columns: {', '.join(missing)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        col_index = {name: idx for idx, name in enumerate(headers)}

        created, skipped = 0, 0
        skipped_words = []
        seen_words = set()

        for row_num, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2
        ):
            word = row[col_index["word"]]

            if not word:
                continue

            word = str(word).strip()

            # 🔹 Optional columns
            meaning_bn = row[col_index["meaning_bn"]] if "meaning_bn" in col_index else ""
            example_en = row[col_index["example_en"]] if "example_en" in col_index else ""
            example_bn = row[col_index["example_bn"]] if "example_bn" in col_index else ""
            hint = row[col_index["hint"]] if "hint" in col_index else ""
            difficulty = row[col_index["difficulty"]] if "difficulty" in col_index else "easy"

            # 🔹 Excel duplicate check
            if word.lower() in seen_words:
                skipped += 1
                skipped_words.append(f"{word} (row {row_num})")
                continue
            seen_words.add(word.lower())

            # 🔹 DB duplicate check
            if Word.objects.filter(puzzle=puzzle, text__iexact=word).exists():
                skipped += 1
                skipped_words.append(f"{word} (already exists)")
                continue

            difficulty = (difficulty or "easy").strip().lower()
            if difficulty not in ["easy", "medium", "hard"]:
                difficulty = "easy"

            Word.objects.create(
                puzzle=puzzle,
                text=word,
                meaning_bn=str(meaning_bn).strip() if meaning_bn else "",
                example_en=str(example_en).strip() if example_en else "",
                example_bn=str(example_bn).strip() if example_bn else "",
                hint=str(hint).strip() if hint else "",
                difficulty=difficulty
            )

            created += 1

        return Response(
            {
                "message": "Excel processed successfully",
                "created": created,
                "skipped": skipped,
                "skipped_words": skipped_words,
            },
            status=status.HTTP_201_CREATED
        )
        

class SubmitWordGameAPIView(APIView):
    """
    Submit a word game attempt.
    - Auth user: uses request.user
    - Guest user: uses username + phone_number
    - Player is UNIQUE
    - Every submit creates a NEW attempt
    """
    permission_classes = [AllowAny]

    def post(self, request):
        puzzle_id = request.data.get("puzzle_id")
        score = request.data.get("score")

        if puzzle_id is None or score is None:
            return Response(
                {"error": "puzzle_id and score are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        puzzle = get_object_or_404(WordPuzzle, id=puzzle_id)

        # ==============================
        # GET OR CREATE PLAYER (SAFE)
        # ==============================
        with transaction.atomic():

            # -------- AUTHENTICATED USER --------
            if request.user and request.user.is_authenticated:
                player = Player.objects.filter(user=request.user).first()

                if not player:
                    player = Player.objects.create(
                        user=request.user,
                        username=request.user.get_username()
                    )

            # -------- GUEST USER --------
            else:
                username = request.data.get("username")
                phone_number = request.data.get("phone_number")

                if not username or not phone_number:
                    return Response(
                        {"error": "username and phone_number are required for guest users"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Reuse player by phone number
                player = Player.objects.filter(phone_number=phone_number).first()

                if not player:
                    player = Player.objects.create(
                        phone_number=phone_number,
                        username=username
                    )

        # ==============================
        # ALWAYS CREATE NEW ATTEMPT
        # ==============================
        attempt = WordGameAttempt.objects.create(
            player=player,
            puzzle=puzzle,
            score=score,
            finished_at=timezone.now()
        )

        return Response(
            {
                "message": "Attempt submitted successfully",
                "attempt_id": attempt.id,
                "player_id": player.id,
                "score": attempt.score
            },
            status=status.HTTP_201_CREATED
        )


class PivotWordGameLeaderboardAPIView(APIView):
    """
    Pivot-style leaderboard:
    Rank | Username | Puzzle1 | Puzzle2 | ... | Total Score
    (SUM of all attempts in date range)
    """
    permission_classes = [AllowAny]

    class StandardResultsSetPagination(PageNumberPagination):
        page_size = 20
        page_size_query_param = "page_size"
        max_page_size = 100

    def get(self, request):
        from_date = request.query_params.get("from_date")
        to_date = request.query_params.get("to_date")

        today = timezone.localdate()
        from_date_parsed = parse_date(from_date) if from_date else today
        to_date_parsed = parse_date(to_date) if to_date else today

        start_datetime = timezone.make_aware(
            datetime.combine(from_date_parsed, time.min)
        )
        end_datetime = timezone.make_aware(
            datetime.combine(to_date_parsed, time.max)
        )

        # All puzzles attempted in date range
        puzzles = WordPuzzle.objects.filter(
            attempts__finished_at__range=(start_datetime, end_datetime)
        ).distinct()

        puzzle_titles = [p.title for p in puzzles]

        # All players who attempted in date range
        players = Player.objects.filter(
            attempts__finished_at__range=(start_datetime, end_datetime)
        ).distinct()

        leaderboard = []

        for player in players:
            print(player.display_name)
            row = {"username": player.display_name}
            total_score = 0

            for puzzle in puzzles:
                # ✅ SUM of ALL attempts for this player & puzzle
                score_sum = WordGameAttempt.objects.filter(
                    player=player,
                    puzzle=puzzle,
                    finished_at__range=(start_datetime, end_datetime)
                ).aggregate(total=Sum("score"))["total"] or 0

                row[puzzle.title] = score_sum
                total_score += score_sum

            row["total_score"] = total_score
            leaderboard.append(row)

        # Sort by total_score DESC
        leaderboard.sort(key=lambda x: x["total_score"], reverse=True)

        # Add rank
        for idx, row in enumerate(leaderboard, start=1):
            row["rank"] = idx

        # Pagination
        paginator = self.StandardResultsSetPagination()
        page = paginator.paginate_queryset(leaderboard, request)

        return paginator.get_paginated_response({
            "puzzle_titles": puzzle_titles,
            "results": page
        })
        
        



class UserGameActivityView(APIView):
    def get(self, request):
        # -----------------------------
        # 1. Identity Logic
        # -----------------------------
        phone = request.query_params.get('phone')
        user = request.user if request.user.is_authenticated else None

        if not user and not phone:
            return Response(
                {"detail": "User authentication or phone number is required"},
                status=400
            )

        lookup_filter = Q(user=user) if user else Q(phone_number=phone)

        # -----------------------------
        # 2. QUIZ (Practice) TOTAL SCORE
        # -----------------------------
        practice_total = (
            PracticeSession.objects
            .filter(lookup_filter)
            .aggregate(total=Sum('score'))
            .get('total') or 0
        )

        # -----------------------------
        # 3. QUIZ RANK (FROM USERPOINTS)
        # SAME LOGIC AS LEADERBOARD
        # -----------------------------
        quiz_rank = None
        user_points = UserPoints.objects.filter(lookup_filter).first()

        if user_points:
            if user:
                rank_list = list(
                    UserPoints.objects
                    .order_by('-points')
                    .values_list('user_id', flat=True)
                )
                try:
                    quiz_rank = rank_list.index(user.id) + 1
                except ValueError:
                    quiz_rank = None
            else:
                rank_list = list(
                    UserPoints.objects
                    .order_by('-points')
                    .values_list('phone_number', flat=True)
                )
                try:
                    quiz_rank = rank_list.index(phone) + 1
                except ValueError:
                    quiz_rank = None

        # -----------------------------
        # 4. GAME / PUZZLE (NO RANK)
        # -----------------------------
        puzzle_total = (
            WordGameAttempt.objects.filter(
                player__user=user if user else None,
                player__phone_number=phone if not user else None
            )
            .aggregate(total=Sum('score'))
            .get('total') or 0
        )

        puzzle_grouped_history = (
            WordGameAttempt.objects.filter(
                player__user=user if user else None,
                player__phone_number=phone if not user else None
            )
            .values('puzzle__title')
            .annotate(total_puzzle_score=Sum('score'))
            .order_by('-total_puzzle_score')
        )

        # -----------------------------
        # 5. Points & Rewards
        # -----------------------------
        rewards_sum = (
            UserReward.objects.filter(
                phone_number=phone or (
                    user.player.phone_number if user and hasattr(user, "player") else ""
                )
            )
            .aggregate(total=Sum('reward_amount'))
            .get('total') or 0.00
        )

        # -----------------------------
        # 6. RESPONSE
        # -----------------------------
        return Response({
            "identity": {
                "phone_number": phone or user.username,
                "access_token_active": bool(user)
            },
            "stats": {
                "total_practice_score": practice_total,
                "quiz_rank": quiz_rank,              # ✅ MATCHES LEADERBOARD
                "total_puzzle_score": puzzle_total,
                "current_points": user_points.points if user_points else 0,
                "total_rewards_taka": rewards_sum
            },
            "history": {
                "practice": PracticeSessionSerializer(
                    PracticeSession.objects
                    .filter(lookup_filter)
                    .order_by('-created_at')[:10],
                    many=True
                ).data,
                "puzzles": list(puzzle_grouped_history)
            }
        })