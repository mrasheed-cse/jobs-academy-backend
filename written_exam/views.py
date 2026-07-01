# exams/api/views.py (or wherever you keep your API views)
import os
import logging
logger = logging.getLogger(__name__)
import uuid
import tempfile
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from PIL import Image
from rest_framework.permissions import AllowAny
from django.core.files.base import ContentFile
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db import transaction # For atomic operations
from django.shortcuts import get_object_or_404
from django.contrib.auth import get_user_model # To get the User model
from rest_framework.permissions import IsAuthenticated
from rest_framework import generics
# Import your specific models from the correct paths
# Assuming your models are in an app called 'exams' and 'quiz'
from .models import RootExam, WrittenExam, Passage, WrittenQuestion, SubWrittenQuestion
from quiz.models import ExamType, Organization, Department, Position, Subject, PastExam
from .serializers import *
User = get_user_model() # Get the currently active user model

class CreateWrittenExamAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        data = request.data
        files = request.FILES
        logger.debug(f"Received data: {data}")
        logger.debug(f"Received files: {files}")

        if not request.user.is_authenticated:
            return Response({"error": "Authentication required."}, status=status.HTTP_401_UNAUTHORIZED)
        
        created_by_user = request.user 

        try:
            with transaction.atomic():
                def get_id_or_none(field_name):
                    value = data.get(field_name)
                    return int(value) if value and str(value).isdigit() else None
                
                exam_type_id = get_id_or_none('exam_type')
                organization_id = get_id_or_none('organization')
                department_id = get_id_or_none('department')
                position_id = get_id_or_none('position')

                exam_type = get_object_or_404(ExamType, id=exam_type_id) if exam_type_id is not None else None
                organization = get_object_or_404(Organization, id=organization_id) if organization_id is not None else None
                department = get_object_or_404(Department, id=department_id) if department_id is not None else None
                position = get_object_or_404(Position, id=position_id) if position_id is not None else None
                
                exam_mode = data.get('exam_mode') 

                if exam_mode not in ['written', 'mcq', 'both']:
                    return Response({"error": "Invalid exam_mode. Must be 'written', 'mcq', or 'both'."}, status=status.HTTP_400_BAD_REQUEST)

                exam_date_str = data.get('exam_date')
                exam_date_obj = None
                if exam_date_str:
                    try:
                        exam_date_obj = date.fromisoformat(exam_date_str)
                    except ValueError:
                        raise ValueError("Invalid exam_date format. Please use 'YYYY-MM-DD'.")

                root_exam = RootExam.objects.create(
                    title=data.get('title'),
                    description=data.get('description', ''),
                    exam_mode=exam_mode,
                    exam_type=exam_type,
                    exam_date=exam_date_obj,
                    organization=organization,
                    department=department,
                    position=position,
                    created_by=created_by_user, 
                )

                subjects_ids = data.getlist('subjects')
                subjects_ids = [sub_id for sub_id in subjects_ids if sub_id and str(sub_id).isdigit()]
                if subjects_ids:
                    subjects = Subject.objects.filter(id__in=subjects_ids)
                    root_exam.subjects.set(subjects)

                past_exam_instance = None
                written_exam_instance = None

                if exam_mode in ['mcq', 'both']:
                    mcq_exam_duration = data.get('mcq_exam_duration') 
                    mcq_exam_pass_mark = data.get('mcq_exam_pass_mark')
                    mcq_exam_negative_mark = data.get('mcq_exam_negative_mark')
                    
                    if not all([mcq_exam_duration, mcq_exam_pass_mark, mcq_exam_negative_mark]):
                        raise ValueError("MCQ duration, pass mark, and negative mark are required when exam_mode is 'mcq' or 'both'.")

                    past_exam_instance = PastExam.objects.create(
                        title=data.get('title'),
                        organization=organization,
                        department=department,
                        position=position,
                        exam_type=exam_type,
                        exam_date=exam_date_obj,
                        duration=mcq_exam_duration,
                        pass_mark=mcq_exam_pass_mark,
                        negative_mark=mcq_exam_negative_mark,
                        created_by=created_by_user,
                        is_published=False
                    )
                    root_exam.past_exam = past_exam_instance
                    root_exam.save()

                    past_exam_instance.refresh_from_db() 

                    if "mcq_file" in files:
                        mcq_file = files["mcq_file"]
                        mcq_processing_result = self.process_questions(mcq_file, past_exam_instance)
                        
                        if mcq_processing_result and "error" in mcq_processing_result:
                            raise ValueError(f"MCQ file processing failed: {mcq_processing_result['error']}")
                        logger.info(f"MCQ questions processed successfully for PastExam ID: {past_exam_instance.id}")
                    else:
                        logger.warning(f"No MCQ file provided for '{exam_mode}' mode. PastExam metadata created without questions.")
                
                # Handle Written exam creation if applicable
                if exam_mode in ['written', 'both']:
                    
                    # Initialize a variable to hold the passage instance if created/selected
                    current_request_passage_instance = None

                    idx = 0
                    has_written_questions = False
                    while True:
                        main_question_text_key = f'main_questions[{idx}][question_text]'
                        if main_question_text_key not in data:
                            break

                        has_written_questions = True

                        if not written_exam_instance:
                            written_exam_subject = root_exam.subjects.first() if root_exam.subjects.exists() else None
                            written_exam_instance = WrittenExam.objects.create(
                                root_exam=root_exam,
                                subject=written_exam_subject, 
                                total_questions=0,
                                total_marks=0.0
                            )
                        
                        main_q_text = data.get(main_question_text_key)
                        main_q_number = data.get(f'main_questions[{idx}][question_number]')
                        try:
                            main_q_marks = float(data.get(f'main_questions[{idx}][question_marks]', 0.0))
                        except (TypeError, ValueError):
                            raise ValueError(f"Invalid marks for main question {idx + 1}")
                        main_q_answer_text = data.get(f'main_questions[{idx}][answer_text]')
                        # Get explanation text/image for main question
                        main_q_explanation_text = data.get(f'main_questions[{idx}][explanation_text]')
                        main_q_explanation_image = files.get(f'main_questions[{idx}][explanation_image]')
                        
                        main_q_image = files.get(f'main_questions[{idx}][question_image]')
                        main_q_answer_image = files.get(f'main_questions[{idx}][answer_image]')

                        written_question_subject = written_exam_instance.subject

                        # --- Handle Passage creation/association for the current main question ---
                        passage_title = data.get(f'main_questions[{idx}][passage][title]')
                        passage_text = data.get(f'main_questions[{idx}][passage][text]')
                        passage_image = files.get(f'main_questions[{idx}][passage][image]')
                        passage_subject_id = data.get(f'main_questions[{idx}][passage][subject]')
                        
                        passage_to_link = None
                        
                        # Only create/update passage if passage data is provided for this specific question
                        if passage_title or passage_text or passage_image:
                            passage_subject = None
                            if passage_subject_id and str(passage_subject_id).isdigit(): 
                                passage_subject = get_object_or_404(Subject, id=int(passage_subject_id))

                            passage_to_link = Passage.objects.create(
                                title=passage_title,
                                text=passage_text,
                                image=passage_image,
                                subject=passage_subject,
                            )
                            # If a new passage is created, make it the 'current_request_passage_instance'
                            # so subsequent questions without their own passage data can link to it.
                            current_request_passage_instance = passage_to_link
                        else:
                            # If no specific passage data for this question, and a 'global' passage exists
                            # from a previous question in this request, link to that.
                            passage_to_link = current_request_passage_instance


                        written_question = WrittenQuestion.objects.create(
                            written_exam=written_exam_instance,
                            subject=written_question_subject, 
                            question_text=main_q_text,
                            question_image=main_q_image,
                            question_number=main_q_number,
                            marks=main_q_marks,
                            answer_text=main_q_answer_text,
                            answer_image=main_q_answer_image,
                            explanation_text=main_q_explanation_text, # Add explanation text
                            explanation_image=main_q_explanation_image, # Add explanation image
                            has_sub_questions=False,
                            passage=passage_to_link # <-- Link the question to the Passage instance
                        )
                        # --- End Passage handling ---

                        # Handle Sub-questions for the current main question
                        sub_idx = 0
                        has_sub_questions_flag = False
                        while True:
                            sub_q_text_key = f'main_questions[{idx}][sub_questions][{sub_idx}][text]'
                            if sub_q_text_key not in data:
                                break

                            has_sub_questions_flag = True

                            sub_q_text = data.get(sub_q_text_key)
                            sub_q_number = data.get(f'main_questions[{idx}][sub_questions][{sub_idx}][number]')
                            try:
                                sub_q_marks = float(data.get(f'main_questions[{idx}][sub_questions][{sub_idx}][marks]', 0.0))
                            except (TypeError, ValueError):
                                raise ValueError(f"Invalid marks for sub-question {sub_idx + 1} of main question {idx + 1}")
                            sub_q_answer_text = data.get(f'main_questions[{idx}][sub_questions][{sub_idx}][answer_text]')
                            # Get explanation text/image for sub-question
                            sub_q_explanation_text = data.get(f'main_questions[{idx}][sub_questions][{sub_idx}][explanation_text]')
                            sub_q_explanation_image = files.get(f'main_questions[{idx}][sub_questions][{sub_idx}][explanation_image]')


                            sub_q_image = files.get(f'main_questions[{idx}][sub_questions][{sub_idx}][image]') 
                            sub_q_answer_image = files.get(f'main_questions[{idx}][sub_questions][{sub_idx}][answer_image]')

                            SubWrittenQuestion.objects.create(
                                parent_question=written_question, 
                                text=sub_q_text,
                                image=sub_q_image, 
                                answer_text=sub_q_answer_text,
                                answer_image=sub_q_answer_image,
                                explanation_text=sub_q_explanation_text, # Add explanation text
                                explanation_image=sub_q_explanation_image, # Add explanation image
                                number=sub_q_number, 
                                marks=sub_q_marks,
                            )
                            sub_idx += 1
                            
                        if has_sub_questions_flag:
                            written_question.has_sub_questions = True
                            written_question.save()

                        idx += 1
                    
                    if has_written_questions and written_exam_instance:
                        written_exam_instance.total_questions = written_exam_instance.questions.count()
                        written_exam_instance.total_marks = sum(q.marks for q in written_exam_instance.questions.all())
                        written_exam_instance.save()
                    elif not has_written_questions and exam_mode == 'written':
                        logger.warning("No written questions provided for 'written' exam mode. WrittenExam might be empty.")

                response_data = {
                    "message": "Exam created successfully!", 
                    "root_exam_id": root_exam.id
                }
                if past_exam_instance:
                    response_data["past_exam_id"] = past_exam_instance.id
                if written_exam_instance:
                    response_data["written_exam_id"] = written_exam_instance.id
                
                return Response(response_data, status=status.HTTP_201_CREATED)

        except ValueError as ve:
            logger.error(f"Validation error during exam creation: {ve}", exc_info=True)
            return Response({"error": str(ve)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.exception(f"An unexpected error occurred during exam creation: {e}")
            return Response({"error": "An unexpected error occurred during exam creation.", "details": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    # --- Re-use the helper methods from PastExamViewSet here ---

    def process_questions(self, file, past_exam):
        try:
            logging.info(f"Starting to process the file for PastExam ID: {past_exam.id}")

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                temp_file.write(file.read())
                temp_file_path = temp_file.name

            try:
                workbook = openpyxl.load_workbook(temp_file_path)
                sheet = workbook.active
                image_map = {}
                for image in sheet._images:
                    if hasattr(image, "anchor") and hasattr(image.anchor, '_from'):
                        cell = image.anchor._from
                        cell_ref = f"{get_column_letter(cell.col + 1)}{cell.row + 1}"
                        image_map[cell_ref] = image
                    else:
                        logging.warning("Found image without a valid anchor.")
                logging.info(f"Found {len(image_map)} images mapped to cells.")
            except Exception as e:
                logging.error(f"Error loading workbook with openpyxl or mapping images: {e}")
                try:
                    os.remove(temp_file_path)
                except OSError:
                    pass
                return {"error": f"Error processing Excel file structure or images: {e}"}

            df = pd.read_excel(temp_file_path, dtype=str)
            df = df.fillna("")

            try:
                os.remove(temp_file_path)
                logging.info(f"Temporary file {temp_file_path} removed.")
            except OSError as e:
                logging.warning(f"Could not remove temporary file {temp_file_path}: {e}")

            df.columns = [str(col).lower().strip() for col in df.columns]

            explanation_column_name = next(
                (col for col in df.columns if col in ["explanation", "ব্যাখ্যা"]),
                None
            )
            
            question_column_name = next((col for col in df.columns if 'question' in col), None)
            if not question_column_name:
                return {"error": "Missing critical column: 'question' column not found."}

            required_columns = [question_column_name]
            has_answer = "answer" in df.columns
            has_difficulty = "difficulty" in df.columns
            has_category = "category" in df.columns

            if has_answer:
                required_columns.append("answer")

            missing_cols = [col for col in required_columns if col not in df.columns]
            if missing_cols:
                return {"error": f"Missing columns: {', '.join(missing_cols)}"}

            option_column_patterns = {
                "option1": "option1", "option2": "option2", "option3": "option3", "option4": "option4",
                "a": "option1", "b": "option2", "c": "option3", "d": "option4",
                "ক": "option1", "খ": "option2", "গ": "option3", "ঘ": "option4",
                "a ": "option1", "b ": "option2", "c ": "option3", "d ": "option4",
                "a.": "option1", "b.": "option2", "c.": "option3", "d.": "option4",
                "1": "option1", "2": "option2", "3": "option3", "4": "option4",
            }

            answer_value_to_normalized_key = {
                "option1": "option1", "option2": "option2", "option3": "option3", "option4": "option4",
                "a": "option1", "b": "option2", "c": "option3", "d": "option4",
                "ক": "option1", "খ": "option2", "গ": "option2", "ঘ": "option4",
                "1": "option1", "2": "option2", "3": "option3", "4": "option4",
                "a ": "option1", "b ": "option2", "c ": "option3", "d ": "option4",
                "a.": "option1", "b.": "option2", "c.": "option3", "d.": "option4",
            }

            detected_option_cols_mapping = {}
            for col in df.columns:
                normalized_col = col.lower().strip().replace(".", "")
                for pattern, normalized_key in option_column_patterns.items():
                    if normalized_col == pattern.lower().strip().replace(".", ""):
                        detected_option_cols_mapping[col] = normalized_key
                        break

            if len(set(detected_option_cols_mapping.values())) < 2:
                logging.warning("Could not detect sufficient option columns (need at least 2 unique options from supported formats).")
                detected_option_cols = []
            else:
                detected_option_cols = sorted([
                    (col, normalized_key) for col, normalized_key in detected_option_cols_mapping.items()
                    if normalized_key in ["option1", "option2", "option3", "option4"]
                ], key=lambda x: x[1])
                logging.info(f"Detected option columns: {detected_option_cols}")

            current_subject = None
            question_count = 0
            updated_questions = 0
            skipped_rows = 0

            # This transaction.atomic() is nested within the outer one, which is fine.
            # It ensures that if processing questions fails, only the question-related
            # changes are rolled back, but since the outer transaction encompasses all,
            # a failure here will roll back the RootExam and PastExam too.
            with transaction.atomic(): 
                last_valid_subject = None
                for index, row in df.iterrows():
                    excel_row_num = index + 2

                    subject_text = str(row.get("subject", "")).strip()
                    if subject_text:
                        current_subject, _ = Subject.objects.get_or_create(name=subject_text)
                        last_valid_subject = current_subject
                    elif last_valid_subject:
                        current_subject = last_valid_subject

                    question_cell = f"{get_column_letter(df.columns.get_loc(question_column_name) + 1)}{excel_row_num}"
                    original_question_text = str(row.get(question_column_name, "")).strip()
                    question_image_data = self.get_image_data_from_map(image_map.get(question_cell))

                    if not original_question_text and not question_image_data:
                        logging.warning(f"Skipping row {excel_row_num}: Question text and image are both missing.")
                        skipped_rows += 1
                        continue

                    question_text_for_db = None if question_image_data else original_question_text

                    category = None
                    category_text = str(row.get("category", "")).strip()
                    if has_category and category_text:
                        category, _ = Category.objects.get_or_create(name=category_text)

                    difficulty_level = None
                    difficulty_text = str(row.get("difficulty", "")).strip()
                    if has_difficulty and difficulty_text:
                        try:
                            difficulty_level = int(difficulty_text)
                        except ValueError:
                            logging.error(f"Invalid difficulty level '{difficulty_text}' at row {excel_row_num}. Setting to None.")
                            difficulty_level = None

                    if question_text_for_db:
                        question, created = Question.objects.get_or_create(
                            text=question_text_for_db,
                            defaults={
                                "subject":current_subject,
                                "marks": 1,
                                "category": category,
                                "difficulty_level": difficulty_level
                            }
                        )
                        if created:
                            question_count += 1
                        else:
                            updated = False
                            if not question.category and category:
                                question.category = category
                                updated = True
                            if question.difficulty_level is None and difficulty_level is not None:
                                question.difficulty_level = difficulty_level
                                updated = True
                            if question.subject != current_subject:
                                question.subject = current_subject
                                updated = True
                            if updated:
                                question.save()
                                updated_questions += 1
                    else:
                        question = Question.objects.create(
                            text=None,
                            subject=current_subject,
                            marks=1,
                            category=category,
                            difficulty_level=difficulty_level
                        )
                        question_count += 1

                    existing_usage = QuestionUsage.objects.filter(
                        question=question,
                        past_exam=past_exam,
                        year=past_exam.exam_date.year
                    ).first()

                    if not existing_usage:
                        QuestionUsage.objects.create(
                            question=question,
                            past_exam=past_exam,
                            year=past_exam.exam_date.year
                        )
                    else:
                        logging.info(f"QuestionUsage already exists for question ID {question.id}, past exam ID {past_exam.id}, and year {past_exam.exam_date.year} at row {excel_row_num}. Skipping creation.")

                    past_exam_question = PastExamQuestion.objects.create(
                        exam=past_exam,
                        question=question,
                        order=index + 1
                    )

                    explanation_text = None
                    explanation_image_data = None

                    if explanation_column_name:
                        explanation_text = str(row.get(explanation_column_name, "")).strip()
                        explanation_cell = f"{get_column_letter(df.columns.get_loc(explanation_column_name) + 1)}{excel_row_num}"
                        explanation_image_data = self.get_image_data_from_map(image_map.get(explanation_cell))

                        if explanation_text:
                            past_exam_question.explanation = explanation_text
                        if explanation_image_data:
                            filename = f"past_explanation_image_{past_exam_question.id}_{uuid.uuid4().hex[:8]}.png"
                            image_file = self.save_image_to_field(explanation_image_data, filename)
                            if image_file:
                                past_exam_question.explanation_image.save(image_file.name, image_file, save=False)

                        past_exam_question.save()
                    
                    if question_image_data:
                        filename = f"question_image_{question.id}_{uuid.uuid4().hex[:8]}.png"
                        image_file = self.save_image_to_field(question_image_data, filename)
                        if image_file:
                            question.image.save(image_file.name, image_file, save=True)
                        else:
                            logging.error(f"Failed to process or save image for question at row {excel_row_num}")
                    
                    options_data_for_answer_check = []

                    if detected_option_cols:
                        for original_option_col, normalized_option_key in detected_option_cols:
                            option_text_original = str(row.get(original_option_col, "")).strip()
                            col_letter = get_column_letter(df.columns.get_loc(original_option_col) + 1)
                            option_cell = f"{col_letter}{excel_row_num}"
                            option_image_data = self.get_image_data_from_map(image_map.get(option_cell))

                            option_text_for_db = None if option_image_data else option_text_original

                            if not option_text_original and not option_image_data:
                                logging.warning(f"Skipping empty option '{original_option_col}' for question at row {excel_row_num}.")
                                continue

                            if option_text_for_db:
                                option_obj, created = QuestionOption.objects.get_or_create(
                                    question=question,
                                    text=option_text_for_db,
                                    defaults={
                                        'is_correct': False
                                    }
                                )
                            else:
                                option_obj = QuestionOption.objects.create(
                                    question=question,
                                    text=None,
                                    is_correct=False
                                )

                            existing_past_exam_question_option = PastExamQuestionOption.objects.filter(
                                question=past_exam_question,
                                option=option_obj
                            ).first()

                            if not existing_past_exam_question_option:
                                PastExamQuestionOption.objects.create(
                                    question=past_exam_question,
                                    option=option_obj
                                )
                            else:
                                logging.info(f"PastExamQuestionOption already exists for option ID {option_obj.id} and PastExamQuestion ID {past_exam_question.id}. Reusing existing link.")

                            if option_image_data:
                                opt_filename = f"option_image_{option_obj.id}_{uuid.uuid4().hex[:8]}.png"
                                opt_image_file = self.save_image_to_field(option_image_data, opt_filename)
                                if opt_image_file:
                                    option_obj.image.save(opt_image_file.name, opt_image_file, save=True)
                                else:
                                    logging.error(f"Failed to process or save image for option '{original_option_col}' at row {excel_row_num}")

                            options_data_for_answer_check.append((normalized_option_key, option_obj))

                    if has_answer and detected_option_cols:
                        correct_answer_raw = str(row.get("answer", "")).strip()
                        correct_answer_normalized = correct_answer_raw.lower().strip().replace(".", "") if correct_answer_raw else None

                        if correct_answer_normalized and correct_answer_normalized in answer_value_to_normalized_key:
                            target_normalized_key = answer_value_to_normalized_key[correct_answer_normalized]
                            found_correct = False
                            for opt_normalized_key, opt_obj in options_data_for_answer_check:
                                if opt_normalized_key == target_normalized_key:
                                    if not opt_obj.is_correct:
                                        opt_obj.is_correct = True
                                        opt_obj.save()
                                        found_correct = True
                                    break
                            if not found_correct:
                                logging.warning(f"Correct answer '{correct_answer_raw}' provided at row {excel_row_num}, but corresponding option '{target_normalized_key}' was not found or processed.")
                        elif correct_answer_raw:
                            logging.warning(f"Could not map correct answer '{correct_answer_raw}' at row {excel_row_num} to a valid option key.")

                past_exam.save()

            return {
                "message": f"Processing complete. {question_count} questions added, {updated_questions} updated, {skipped_rows} rows skipped.",
                "past_exam_id": past_exam.id
            }
        except Exception as e:
            logging.exception(f"Critical error during question processing: {str(e)}")
            if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except OSError:
                    pass
            return {"error": f"An unexpected error occurred during MCQ processing: {str(e)}"}

    def get_image_data_from_map(self, image):
        if not image:
            return None
        try:
            img_byte_arr = BytesIO()
            if hasattr(image, '_data'):
                data = image._data()
                img_byte_arr.write(data)
                return img_byte_arr.getvalue()
            else:
                print(f"Error extracting image from cell: Unknown -> Image object has no '_data' attribute")
                return None
        except Exception as e:
            print(f"Error extracting image from cell: {getattr(image.anchor, 'from', 'Unknown')} -> {e}")
            return None

    def save_image_to_field(self, image_data, filename):
        try:
            img = Image.open(BytesIO(image_data))
            if img.mode == 'P' and 'transparency' in img.info:
                img = img.convert('RGBA')
            elif img.mode == 'RGBA':
                pass 
            elif img.mode != 'RGB':
                img = img.convert('RGB')

            img_format = img.format or "PNG"
            img_io = BytesIO()
            img.save(img_io, format=img_format)
            img_io.seek(0)
            return ContentFile(img_io.getvalue(), name=filename)
        except Exception as e:
            logging.error(f"Error processing or saving image '{filename}': {e}")
            return None




# class CreateWrittenExamAPIView(APIView):
#     permission_classes = [IsAuthenticated]

#     def post(self, request, *args, **kwargs):
#         data = request.data
#         files = request.FILES
        
#         if not request.user.is_authenticated:
#             return Response({"error": "Authentication required."}, status=status.HTTP_401_UNAUTHORIZED)
        
#         created_by_user = request.user 

#         try:
#             with transaction.atomic():
#                 exam_type = get_object_or_404(ExamType, id=data.get('exam_type')) if data.get('exam_type') else None
#                 organization = get_object_or_404(Organization, id=data.get('organization')) if data.get('organization') else None
#                 department = get_object_or_404(Department, id=data.get('department')) if data.get('department') else None
#                 position = get_object_or_404(Position, id=data.get('position')) if data.get('position') else None
                
#                 exam_mode = data.get('exam_mode') # Get exam_mode here

#                 root_exam = RootExam.objects.create(
#                     title=data.get('title'),
#                     description=data.get('description', ''),
#                     exam_mode=exam_mode, # Use the retrieved exam_mode
#                     exam_type=exam_type,
#                     exam_date=data.get('exam_date'),
#                     organization=organization,
#                     department=department,
#                     position=position,
#                     created_by=created_by_user, 
#                 )

#                 subjects_ids = data.getlist('subjects')
#                 if subjects_ids:
#                     subjects = Subject.objects.filter(id__in=subjects_ids)
#                     root_exam.subjects.set(subjects)

#                 # --- NEW: Create PastExam if exam_mode is 'both' ---
#                 if exam_mode == 'both':
#                     # Extract PastExam specific fields from request data
#                     # These fields would need to be sent from the frontend if exam_mode is 'both'
#                     # The frontend currently only sends these for 'mcq' mode.
#                     # You would need to add these fields to your form/JS for 'both' mode if they are required.
#                     past_exam_duration = data.get('mcq_exam_duration') 
#                     past_exam_pass_mark = data.get('mcq_exam_pass_mark')
#                     past_exam_negative_mark = data.get('mcq_exam_negative_mark')
                    
#                     past_exam_instance = PastExam.objects.create(
#                         title=data.get('title'), # Re-use RootExam title
#                         organization=organization,
#                         department=department,
#                         position=position,
#                         exam_type=exam_type,
#                         exam_date=data.get('exam_date'),
#                         duration=past_exam_duration,
#                         pass_mark=past_exam_pass_mark,
#                         negative_mark=past_exam_negative_mark,
#                         # Add other PastExam fields as necessary, e.g., created_by
#                         # You might also need to handle the 'file' if 'both' means
#                         # a written exam with an *optional* MCQ file upload.
#                         # However, your JS currently sends 'file' only for 'mcq' mode.
#                         # If 'both' means "written questions + an MCQ excel", then
#                         # the 'file' field from the frontend would need to be present
#                         # in the request.FILES for this view too.
#                         # For now, let's assume 'both' means 'written questions' primarily
#                         # with the *metadata* of a past exam.
#                         created_by=created_by_user,
#                     )
#                     root_exam.past_exam = past_exam_instance
#                     root_exam.save() # Save root_exam to update the past_exam link
#                 # --- END NEW ---

#                 # --- 2. Process Main Questions and their related objects ---
#                 # ... (rest of your existing logic for WrittenExam, WrittenQuestion, Passage, SubWrittenQuestion) ...
                
#                 # IMPORTANT: Your current loop structure for WrittenExam implies one WrittenExam 
#                 # for each main_question. If a RootExam should have *one* WrittenExam
#                 # instance that aggregates all main questions, you'd need to create 
#                 # `written_exam_instance` *before* the loop and then update its 
#                 # total_questions/marks after the loop.
#                 # Assuming your current structure means a WrittenExam per main question:

#                 idx = 0
#                 while True:
#                     main_question_text_key = f'main_questions[{idx}][question_text]'
#                     if main_question_text_key not in data:
#                         break

#                     # Fetch subject for WrittenExam. If RootExam has multiple subjects, 
#                     # you might need a way to determine which subject applies to this specific WrittenExam instance.
#                     # For now, taking the first subject of the root_exam or making it optional.
#                     # If a WrittenExam can have multiple subjects, your model would need a ManyToMany.
#                     # Current model: subject = ForeignKey(Subject, on_delete=models.CASCADE)
#                     # So, one subject per WrittenExam.
#                     written_exam_subject = root_exam.subjects.first() if root_exam.subjects.exists() else None
#                     # You might consider letting the user pick the subject for each WrittenExam instance
#                     # if the RootExam has multiple subjects.
                    
#                     # Create WrittenExam instance for this group of questions
#                     written_exam_instance = WrittenExam.objects.create(
#                         root_exam=root_exam,
#                         subject=written_exam_subject, 
#                         total_questions=0, # Will be updated after processing main and sub questions
#                         total_marks=0.0    # Will be updated after processing main and sub questions
#                     )

#                     current_written_exam_total_questions = 0
#                     current_written_exam_total_marks = 0.0

#                     main_q_text = data.get(main_question_text_key)
#                     main_q_number = data.get(f'main_questions[{idx}][question_number]')
#                     main_q_marks = float(data.get(f'main_questions[{idx}][question_marks]', 0.0))
#                     main_q_answer_text = data.get(f'main_questions[{idx}][answer_text]')
                    
#                     main_q_image = files.get(f'main_questions[{idx}][question_image]')
#                     main_q_answer_image = files.get(f'main_questions[{idx}][answer_image]')

#                     # Subject for WrittenQuestion can be derived from the WrittenExam's subject
#                     # or left null if it's truly optional per question.
#                     # Based on your model `subject = models.ForeignKey(Subject, on_delete=models.SET_NULL, blank=True, null=True)`
#                     written_question_subject = written_exam_subject # Or provide a way to select per question
                    
#                     written_question = WrittenQuestion.objects.create(
#                         written_exam=written_exam_instance,
#                         subject=written_question_subject, 
#                         question_text=main_q_text,
#                         question_image=main_q_image,
#                         question_number=main_q_number,
#                         marks=main_q_marks,
#                         answer_text=main_q_answer_text,
#                         answer_image=main_q_answer_image,
#                         has_sub_questions=False # Default to False, set to True if sub-questions found
#                     )
                    
#                     current_written_exam_total_questions += 1
#                     current_written_exam_total_marks += main_q_marks

#                     passage_title_key = f'main_questions[{idx}][passage][title]'
#                     if data.get(f'main_questions[{idx}][passage][text]') or files.get(f'main_questions[{idx}][passage][image]'): # Check if passage content exists
#                         passage_subject_id = data.get(f'main_questions[{idx}][passage][subject]')
#                         passage_subject = get_object_or_404(Subject, id=passage_subject_id) if passage_subject_id else None
                        
#                         Passage.objects.create(
#                             written_question=written_question,
#                             title=data.get(passage_title_key),
#                             text=data.get(f'main_questions[{idx}][passage][text]'),
#                             image=files.get(f'main_questions[{idx}][passage][image]'),
#                             subject=passage_subject,
#                         )

#                     sub_idx = 0
#                     has_sub_questions_flag = False
#                     while True:
#                         sub_q_text_key = f'main_questions[{idx}][sub_questions][{sub_idx}][text]'
#                         if sub_q_text_key not in data:
#                             break

#                         has_sub_questions_flag = True

#                         sub_q_text = data.get(sub_q_text_key)
#                         sub_q_number = data.get(f'main_questions[{idx}][sub_questions][{sub_idx}][number]')
#                         sub_q_marks = float(data.get(f'main_questions[{idx}][sub_questions][{sub_idx}][marks]', 0.0))
#                         sub_q_answer_text = data.get(f'main_questions[{idx}][sub_questions][{sub_idx}][answer_text]')

#                         sub_q_image = files.get(f'main_questions[{idx}][sub_questions][{sub_idx}][image]') 
#                         sub_q_answer_image = files.get(f'main_questions[{idx}][sub_questions][{sub_idx}][answer_image]')

#                         SubWrittenQuestion.objects.create(
#                             parent_question=written_question, 
#                             text=sub_q_text,
#                             image=sub_q_image, 
#                             answer_text=sub_q_answer_text,
#                             answer_image=sub_q_answer_image,
#                             number=sub_q_number, 
#                             marks=sub_q_marks,
#                         )
                        
#                         current_written_exam_total_questions += 1
#                         current_written_exam_total_marks += sub_q_marks
#                         sub_idx += 1
                    
#                     if has_sub_questions_flag:
#                         written_question.has_sub_questions = True
#                         written_question.save()

#                     # Update the current WrittenExam instance after processing all its main and sub questions
#                     written_exam_instance.total_questions = current_written_exam_total_questions
#                     written_exam_instance.total_marks = current_written_exam_total_marks
#                     written_exam_instance.save()
                    
#                     idx += 1

#                 return Response({"message": "Written exam created successfully!", "exam_id": root_exam.id}, status=status.HTTP_201_CREATED)

#         except Exception as e:
#             print(f"Error creating written exam: {e}")
#             return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)




# class RootExamListView(generics.ListAPIView):
    
#     serializer_class = RootExamSerializer
#     # permission_classes = [IsAuthenticated] # <--- REMOVE OR COMMENT OUT THIS LINE
#     permission_classes = [] # Allow any access

#     def get_queryset(self):
#         """
#         Filters the queryset based on the user's role,
#         but now also accessible to unauthenticated users.
#         """
#         user = self.request.user

#         # Handle unauthenticated users: They won't have a role,
#         # so decide what they should see.
#         if not user.is_authenticated:
#             # For example, let unauthenticated users see all exams.
#             # Or you could return RootExam.objects.none() if they should see nothing.
#             return RootExam.objects.all().order_by('-created_at') # 🚨 Security Consideration Here!

#         # Existing logic for authenticated users remains the same
#         if user.role == User.ADMIN or user.role == User.STUDENT:
#             return RootExam.objects.all().order_by('-created_at')

#         elif user.role == User.TEACHER or user.role == User.OPERATOR:
#             return RootExam.objects.filter(created_by=user).order_by('-created_at')

#         return RootExam.objects.none()



class RootExamListView(generics.ListAPIView):
    serializer_class = RootExamSerializer
    permission_classes = [AllowAny]  # Allow any access

    def get_queryset(self):
        """
        Filters the queryset based on exam type, and additionally by user role.
        """
        user = self.request.user
        
        # Get the exam_type_id from the URL query parameters
        exam_type_id = self.request.query_params.get('exam_type')

        # Start with the base queryset
        queryset = RootExam.objects.all().order_by('-created_at')

        # Always filter by exam_type if the parameter is provided
        if exam_type_id:
            queryset = queryset.filter(exam_type__id=exam_type_id)

        # Apply additional filters based on user role if the user is authenticated
        if user.is_authenticated:
            if user.role in [user.ADMIN, user.STUDENT]:
                # ADMIN and STUDENT can see all exams of the specified type
                pass
            
            elif user.role in [user.TEACHER, user.OPERATOR]:
                # TEACHER and OPERATOR can only see exams they created of the specified type
                queryset = queryset.filter(created_by=user)
        
        # Prefetch related data to optimize performance and prevent N+1 queries
        queryset = queryset.prefetch_related('written_exams', 'exam_type', 'subjects', 'organization', 'department', 'position')
        
        return queryset

    
    
class RootExamDetailView(generics.RetrieveAPIView):
    """
    API view to retrieve a single RootExam instance by its ID.
    """
    queryset = RootExam.objects.all() # The queryset from which to retrieve the object
    serializer_class = RootExamSerializer
    # permission_classes = [IsAuthenticated] 
    lookup_field = 'pk' # Or 'id' if you prefer, but 'pk' is conventional for primary key
